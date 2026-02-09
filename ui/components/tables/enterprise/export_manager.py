"""
Export Manager
==============
Data export manager (Excel, PDF, CSV). Styling handled by centralized theme system.
"""

from PyQt5.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QRadioButton,
    QPushButton, QLabel, QFileDialog, QCheckBox,
    QGroupBox, QButtonGroup, QProgressBar, QMessageBox
)
from PyQt5.QtCore import pyqtSignal, QThread

from core.themes import get_current_palette, get_font, FONT_SIZE_SUBTITLE, FONT_WEIGHT_BOLD

import os
from datetime import datetime


class ExportWorker(QThread):
    """Worker thread for export operations."""

    progress = pyqtSignal(int)
    finished = pyqtSignal(bool, str)

    def __init__(self, data, columns, filepath, export_format, include_headers=True):
        super().__init__()
        self._data = data
        self._columns = columns
        self._filepath = filepath
        self._format = export_format
        self._include_headers = include_headers

    def run(self):
        """Run export in background."""
        try:
            if self._format == 'excel':
                self._export_excel()
            elif self._format == 'csv':
                self._export_csv()
            elif self._format == 'pdf':
                self._export_pdf()

            self.finished.emit(True, self._filepath)
        except Exception as e:
            self.finished.emit(False, str(e))

    def _export_excel(self):
        """Export to Excel."""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            raise Exception("\u0645\u0643\u062a\u0628\u0629 openpyxl \u063a\u064a\u0631 \u0645\u062b\u0628\u062a\u0629")

        palette = get_current_palette()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "\u0627\u0644\u0628\u064a\u0627\u0646\u0627\u062a"

        # RTL support
        ws.sheet_view.rightToLeft = True

        # Header style - use palette primary color
        primary_hex = palette['primary'].lstrip('#')
        header_fill = PatternFill(start_color=primary_hex, end_color=primary_hex, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal='center', vertical='center')

        # Cell style
        cell_alignment = Alignment(horizontal='center', vertical='center')
        border_hex = palette['border'].lstrip('#')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write headers
        data_start_row = 1
        if self._include_headers:
            for col, header in enumerate(self._columns, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
            data_start_row = 2

        # Write data
        total = len(self._data)
        for row_idx, row_data in enumerate(self._data, data_start_row):
            if isinstance(row_data, dict):
                values = [row_data.get(col, "") for col in self._columns]
            else:
                values = list(row_data)

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border

            # Progress
            if total > 0:
                progress = int((row_idx / total) * 100)
                self.progress.emit(progress)

        # Auto-fit columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        wb.save(self._filepath)

    def _export_csv(self):
        """Export to CSV."""
        import csv

        with open(self._filepath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)

            # Write header
            if self._include_headers:
                writer.writerow(self._columns)

            # Write data
            total = len(self._data)
            for idx, row_data in enumerate(self._data):
                if isinstance(row_data, dict):
                    values = [row_data.get(col, "") for col in self._columns]
                else:
                    values = list(row_data)

                writer.writerow(values)

                # Progress
                if total > 0:
                    progress = int(((idx + 1) / total) * 100)
                    self.progress.emit(progress)

    def _export_pdf(self):
        """Export to PDF."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
        except ImportError:
            raise Exception("\u0645\u0643\u062a\u0628\u0629 reportlab \u063a\u064a\u0631 \u0645\u062b\u0628\u062a\u0629")

        palette = get_current_palette()

        # Create PDF
        doc = SimpleDocTemplate(
            self._filepath,
            pagesize=landscape(A4),
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )

        elements = []

        # Prepare data
        table_data = [self._columns] if self._include_headers else []

        total = len(self._data)
        for idx, row_data in enumerate(self._data):
            if isinstance(row_data, dict):
                values = [str(row_data.get(col, "")) if row_data.get(col) else "" for col in self._columns]
            else:
                values = [str(v) if v else "" for v in row_data]

            table_data.append(values)

            # Progress
            if total > 0:
                progress = int(((idx + 1) / total) * 100)
                self.progress.emit(progress)

        # Create table
        table = Table(table_data, repeatRows=1)

        # Table style - use palette colors
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(palette['primary'])),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor(palette['border'])),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor(palette['bg_main'])]),
        ])
        table.setStyle(style)

        elements.append(table)
        doc.build(elements)


class ExportManager(QDialog):
    """
    Export dialog for table data.
    Styling handled by centralized theme system.

    Supports:
    - Excel (.xlsx)
    - CSV (.csv)
    - PDF (.pdf)
    """

    export_started = pyqtSignal()
    export_finished = pyqtSignal(bool, str)

    def __init__(self, data: list, columns: list, parent=None):
        super().__init__(parent)

        self._data = data
        self._columns = columns
        self._worker = None

        self._setup_ui()
        # App-level QSS handles dialog, button, radio, checkbox, progress styling

    def _setup_ui(self):
        """Setup dialog UI."""
        self.setWindowTitle("\U0001f4e4 \u062a\u0635\u062f\u064a\u0631 \u0627\u0644\u0628\u064a\u0627\u0646\u0627\u062a")
        self.setMinimumSize(400, 300)
        self.setModal(True)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # Title
        title = QLabel(f"\u062a\u0635\u062f\u064a\u0631 {len(self._data)} \u0633\u062c\u0644")
        title.setFont(get_font(FONT_SIZE_SUBTITLE, FONT_WEIGHT_BOLD))
        layout.addWidget(title)

        # Format selection
        format_group = QGroupBox("\u0627\u062e\u062a\u0631 \u0635\u064a\u063a\u0629 \u0627\u0644\u062a\u0635\u062f\u064a\u0631:")
        format_layout = QVBoxLayout(format_group)

        self._format_group = QButtonGroup(self)

        self._excel_radio = QRadioButton("\U0001f4ca Excel (.xlsx)")
        self._excel_radio.setChecked(True)
        self._format_group.addButton(self._excel_radio)
        format_layout.addWidget(self._excel_radio)

        self._csv_radio = QRadioButton("\U0001f4c4 CSV (.csv)")
        self._format_group.addButton(self._csv_radio)
        format_layout.addWidget(self._csv_radio)

        self._pdf_radio = QRadioButton("\U0001f4d5 PDF (.pdf)")
        self._format_group.addButton(self._pdf_radio)
        format_layout.addWidget(self._pdf_radio)

        layout.addWidget(format_group)

        # Options
        self._include_headers = QCheckBox("\u062a\u0636\u0645\u064a\u0646 \u0627\u0644\u0639\u0646\u0627\u0648\u064a\u0646")
        self._include_headers.setChecked(True)
        layout.addWidget(self._include_headers)

        # Progress bar
        self._progress = QProgressBar()
        self._progress.setVisible(False)
        layout.addWidget(self._progress)

        # Status label
        self._status_label = QLabel("")
        self._status_label.setVisible(False)
        layout.addWidget(self._status_label)

        layout.addStretch()

        # Buttons
        buttons_layout = QHBoxLayout()
        buttons_layout.addStretch()

        cancel_btn = QPushButton("\u0625\u0644\u063a\u0627\u0621")
        cancel_btn.setProperty("cssClass", "secondary")
        cancel_btn.clicked.connect(self.reject)
        buttons_layout.addWidget(cancel_btn)

        self._export_btn = QPushButton("\U0001f4e4 \u062a\u0635\u062f\u064a\u0631")
        self._export_btn.clicked.connect(self._start_export)
        buttons_layout.addWidget(self._export_btn)

        layout.addLayout(buttons_layout)

    def _start_export(self):
        """Start export process."""
        # Get format
        if self._excel_radio.isChecked():
            export_format = 'excel'
            file_filter = "Excel Files (*.xlsx)"
            default_ext = ".xlsx"
        elif self._csv_radio.isChecked():
            export_format = 'csv'
            file_filter = "CSV Files (*.csv)"
            default_ext = ".csv"
        else:
            export_format = 'pdf'
            file_filter = "PDF Files (*.pdf)"
            default_ext = ".pdf"

        # Get save path
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"export_{timestamp}{default_ext}"

        filepath, _ = QFileDialog.getSaveFileName(
            self,
            "\u062d\u0641\u0638 \u0627\u0644\u0645\u0644\u0641",
            default_name,
            file_filter
        )

        if not filepath:
            return

        # Show progress
        self._progress.setVisible(True)
        self._progress.setValue(0)
        self._status_label.setVisible(True)
        self._status_label.setText("\u062c\u0627\u0631\u064a \u0627\u0644\u062a\u0635\u062f\u064a\u0631...")
        self._export_btn.setEnabled(False)

        # Start worker
        include_headers = self._include_headers.isChecked()
        self._worker = ExportWorker(self._data, self._columns, filepath, export_format, include_headers)
        self._worker.progress.connect(self._on_progress)
        self._worker.finished.connect(self._on_finished)
        self._worker.start()

        self.export_started.emit()

    def _on_progress(self, value: int):
        """Handle progress update."""
        self._progress.setValue(value)

    def _on_finished(self, success: bool, message: str):
        """Handle export finished."""
        self._export_btn.setEnabled(True)

        if success:
            self._status_label.setText("\u2705 \u062a\u0645 \u0627\u0644\u062a\u0635\u062f\u064a\u0631 \u0628\u0646\u062c\u0627\u062d!")
            self.export_finished.emit(True, message)

            # Ask to open file
            reply = QMessageBox.question(
                self,
                "\u062a\u0645 \u0627\u0644\u062a\u0635\u062f\u064a\u0631",
                "\u062a\u0645 \u062a\u0635\u062f\u064a\u0631 \u0627\u0644\u0628\u064a\u0627\u0646\u0627\u062a \u0628\u0646\u062c\u0627\u062d.\n\u0647\u0644 \u062a\u0631\u064a\u062f \u0641\u062a\u062d \u0627\u0644\u0645\u0644\u0641\u061f",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes
            )

            if reply == QMessageBox.Yes:
                import sys
                import subprocess
                if sys.platform == 'win32':
                    os.startfile(message)
                elif sys.platform == 'darwin':
                    subprocess.Popen(['open', message])
                else:
                    subprocess.Popen(['xdg-open', message])

            self.accept()
        else:
            self._status_label.setText(f"\u274c \u062e\u0637\u0623: {message}")
            self.export_finished.emit(False, message)
