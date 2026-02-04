# -*- coding: utf-8 -*-
"""
BI Data Exporter
================
Export BI Views data to CSV and Excel formats for Power BI.

This module provides:
- Export to CSV with Arabic support
- Export to Excel with multiple sheets
- Batch export of all views
- Export history tracking

Author: Mohamed
Version: 1.0.0
Date: February 2026
"""

import os
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional, Tuple
from dataclasses import dataclass

from core.database import select_all
from core.logging import app_logger


@dataclass
class ExportResult:
    """Result of an export operation."""
    success: bool
    file_path: str
    view_name: str
    row_count: int
    file_size: int  # in bytes
    export_time: datetime
    error: str = ""


class BIDataExporter:
    """
    Exports BI Views data to various formats.

    Supports CSV and Excel export with proper encoding
    for Arabic text compatibility.
    """

    def __init__(self, export_path: Optional[str] = None):
        """
        Initialize the exporter.

        Args:
            export_path: Custom export directory path
        """
        from .connection_config import get_export_path, BI_EXPORT_CONFIG

        self._export_path = Path(export_path) if export_path else get_export_path()
        self._config = BI_EXPORT_CONFIG
        self._export_history: List[ExportResult] = []

        # Ensure export directory exists
        self._export_path.mkdir(parents=True, exist_ok=True)

    @property
    def export_path(self) -> Path:
        """Get the export directory path."""
        return self._export_path

    def export_to_csv(
        self,
        view_name: str,
        output_path: Optional[str] = None,
        encoding: str = "utf-8-sig",
        delimiter: str = ","
    ) -> ExportResult:
        """
        Export a BI View to CSV file.

        Args:
            view_name: Name of the view to export
            output_path: Custom output file path
            encoding: File encoding (utf-8-sig for Excel Arabic support)
            delimiter: CSV delimiter character

        Returns:
            ExportResult with operation details
        """
        timestamp = datetime.now()
        filename = f"{view_name}_{timestamp.strftime('%Y%m%d_%H%M%S')}.csv"
        file_path = output_path or str(self._export_path / filename)

        try:
            # Fetch data from view
            sql = f"SELECT * FROM bi_views.{view_name}"
            columns, rows = select_all(sql)

            if not columns:
                return ExportResult(
                    success=False,
                    file_path=file_path,
                    view_name=view_name,
                    row_count=0,
                    file_size=0,
                    export_time=timestamp,
                    error="No data or view not found"
                )

            # Write CSV file
            with open(file_path, 'w', encoding=encoding, newline='') as f:
                # Write header
                f.write(delimiter.join(columns) + '\n')

                # Write data rows
                for row in rows:
                    # Handle None values and format data
                    formatted = []
                    for val in row:
                        if val is None:
                            formatted.append('')
                        elif isinstance(val, (int, float)):
                            formatted.append(str(val))
                        elif isinstance(val, datetime):
                            formatted.append(val.strftime('%Y-%m-%d %H:%M:%S'))
                        else:
                            # Escape quotes and handle special characters
                            str_val = str(val).replace('"', '""')
                            if delimiter in str_val or '"' in str_val or '\n' in str_val:
                                str_val = f'"{str_val}"'
                            formatted.append(str_val)
                    f.write(delimiter.join(formatted) + '\n')

            # Get file size
            file_size = os.path.getsize(file_path)

            result = ExportResult(
                success=True,
                file_path=file_path,
                view_name=view_name,
                row_count=len(rows),
                file_size=file_size,
                export_time=timestamp
            )

            app_logger.info(f"Exported {view_name} to CSV: {len(rows)} rows")
            self._export_history.append(result)
            return result

        except Exception as e:
            error_msg = str(e)
            app_logger.error(f"Failed to export {view_name} to CSV: {error_msg}")

            return ExportResult(
                success=False,
                file_path=file_path,
                view_name=view_name,
                row_count=0,
                file_size=0,
                export_time=timestamp,
                error=error_msg
            )

    def export_to_excel(
        self,
        views: List[str],
        output_path: Optional[str] = None,
        include_metadata: bool = True
    ) -> ExportResult:
        """
        Export multiple BI Views to a single Excel file.

        Each view becomes a separate sheet in the workbook.

        Args:
            views: List of view names to export
            output_path: Custom output file path
            include_metadata: Include a summary sheet

        Returns:
            ExportResult with operation details
        """
        timestamp = datetime.now()
        filename = f"bi_export_{timestamp.strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = output_path or str(self._export_path / filename)

        try:
            # Try to import openpyxl
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                from openpyxl.utils import get_column_letter
            except ImportError:
                return ExportResult(
                    success=False,
                    file_path=file_path,
                    view_name="multiple",
                    row_count=0,
                    file_size=0,
                    export_time=timestamp,
                    error="openpyxl is not installed. Run: pip install openpyxl"
                )

            wb = Workbook()
            total_rows = 0

            # Style definitions
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

            # Export each view to a sheet
            for view_name in views:
                try:
                    sql = f"SELECT * FROM bi_views.{view_name}"
                    columns, rows = select_all(sql)

                    if not columns:
                        continue

                    # Create sheet (truncate name to 31 chars for Excel limit)
                    sheet_name = view_name[:31]
                    ws = wb.create_sheet(title=sheet_name)

                    # Write header row with styling
                    for col_idx, col_name in enumerate(columns, 1):
                        cell = ws.cell(row=1, column=col_idx, value=col_name)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment

                    # Write data rows
                    for row_idx, row in enumerate(rows, 2):
                        for col_idx, val in enumerate(row, 1):
                            if isinstance(val, datetime):
                                ws.cell(row=row_idx, column=col_idx, value=val)
                            else:
                                ws.cell(row=row_idx, column=col_idx, value=val)

                    # Auto-adjust column widths
                    for col_idx, col_name in enumerate(columns, 1):
                        max_length = len(str(col_name))
                        for row in rows[:100]:  # Sample first 100 rows
                            cell_value = str(row[col_idx - 1]) if row[col_idx - 1] else ""
                            max_length = max(max_length, len(cell_value))
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

                    # Freeze header row
                    ws.freeze_panes = "A2"

                    total_rows += len(rows)

                except Exception as e:
                    app_logger.warning(f"Failed to export view {view_name}: {e}")
                    continue

            # Add metadata sheet if requested
            if include_metadata:
                ws_meta = wb.create_sheet(title="Export Info", index=0)
                ws_meta["A1"] = "Export Date"
                ws_meta["B1"] = timestamp.strftime('%Y-%m-%d %H:%M:%S')
                ws_meta["A2"] = "Views Exported"
                ws_meta["B2"] = ", ".join(views)
                ws_meta["A3"] = "Total Rows"
                ws_meta["B3"] = total_rows
                ws_meta["A4"] = "Generated By"
                ws_meta["B4"] = "INTEGRA BI Module"

            # Save workbook
            wb.save(file_path)
            file_size = os.path.getsize(file_path)

            result = ExportResult(
                success=True,
                file_path=file_path,
                view_name="multiple",
                row_count=total_rows,
                file_size=file_size,
                export_time=timestamp
            )

            app_logger.info(f"Exported {len(views)} views to Excel: {total_rows} total rows")
            self._export_history.append(result)
            return result

        except Exception as e:
            error_msg = str(e)
            app_logger.error(f"Failed to export to Excel: {error_msg}")

            return ExportResult(
                success=False,
                file_path=file_path,
                view_name="multiple",
                row_count=0,
                file_size=0,
                export_time=timestamp,
                error=error_msg
            )

    def export_all_views_csv(self) -> List[ExportResult]:
        """Export all BI Views to individual CSV files."""
        from .views_manager import SQL_VIEWS

        results = []
        for view_name in SQL_VIEWS.keys():
            result = self.export_to_csv(view_name)
            results.append(result)
        return results

    def export_all_views_excel(self, output_path: Optional[str] = None) -> ExportResult:
        """Export all BI Views to a single Excel file."""
        from .views_manager import SQL_VIEWS

        return self.export_to_excel(
            views=list(SQL_VIEWS.keys()),
            output_path=output_path,
            include_metadata=True
        )

    def get_export_history(self) -> List[ExportResult]:
        """Get the history of export operations."""
        return self._export_history.copy()

    def clear_history(self) -> None:
        """Clear the export history."""
        self._export_history.clear()

    def cleanup_old_exports(self, retention_days: int = 30) -> int:
        """
        Remove export files older than retention period.

        Args:
            retention_days: Number of days to keep files

        Returns:
            Number of files deleted
        """
        from datetime import timedelta

        cutoff_date = datetime.now() - timedelta(days=retention_days)
        deleted_count = 0

        for file_path in self._export_path.glob("*"):
            if file_path.is_file():
                file_time = datetime.fromtimestamp(file_path.stat().st_mtime)
                if file_time < cutoff_date:
                    try:
                        file_path.unlink()
                        deleted_count += 1
                    except Exception:
                        pass

        if deleted_count > 0:
            app_logger.info(f"Cleaned up {deleted_count} old export files")

        return deleted_count


# =============================================================================
# Singleton Instance
# =============================================================================

_exporter_instance: Optional[BIDataExporter] = None


def get_bi_exporter() -> BIDataExporter:
    """Get the singleton BIDataExporter instance."""
    global _exporter_instance
    if _exporter_instance is None:
        _exporter_instance = BIDataExporter()
    return _exporter_instance
