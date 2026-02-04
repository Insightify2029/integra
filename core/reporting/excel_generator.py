"""
Excel Generator
===============
Generate professional Excel reports using openpyxl.

Features:
- Multiple sheets support
- Table formatting with styles
- Charts (bar, pie, line, etc.)
- RTL and Arabic support
- Formulas and calculations
- Cell formatting
- Data validation
- Conditional formatting

Usage:
    from core.reporting import ExcelGenerator, ExcelConfig

    # Basic usage
    excel = ExcelGenerator()
    excel.add_sheet("الموظفين", data, headers=["الاسم", "الراتب"])
    excel.save("report.xlsx")

    # With charts
    excel = ExcelGenerator()
    excel.add_sheet("المبيعات", sales_data)
    excel.add_chart("المبيعات", chart_type="bar", data_range="A1:B10")
    excel.save("sales_report.xlsx")
"""

from dataclasses import dataclass, field
from typing import List, Dict, Any, Optional, Tuple, Union
from pathlib import Path
from datetime import datetime
import io

from core.logging import app_logger


# Check openpyxl availability
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import (
        Font, Alignment, Border, Side, PatternFill,
        NamedStyle, numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import (
        BarChart, PieChart, LineChart, AreaChart, DoughnutChart,
        Reference, Series
    )
    from openpyxl.chart.label import DataLabelList
    from openpyxl.formatting.rule import (
        ColorScaleRule, FormulaRule, CellIsRule
    )
    from openpyxl.worksheet.datavalidation import DataValidation
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


@dataclass
class ExcelConfig:
    """Excel generation configuration."""
    # Document info
    title: str = ""
    author: str = "INTEGRA"
    company: str = "INTEGRA"
    subject: str = ""

    # Styling
    rtl: bool = True
    font_family: str = "Cairo"
    font_size: int = 11
    header_bg_color: str = "2563eb"
    header_fg_color: str = "ffffff"
    alt_row_color: str = "f5f5f5"
    border_color: str = "dddddd"

    # Layout
    freeze_header: bool = True
    auto_filter: bool = True
    auto_width: bool = True

    # Protection
    protect_sheet: bool = False
    protect_password: Optional[str] = None


class ExcelGenerator:
    """
    Generate Excel reports using openpyxl.

    Supports multiple sheets, charts, and professional formatting.
    """

    def __init__(self, config: ExcelConfig = None):
        """
        Initialize Excel generator.

        Args:
            config: Excel configuration
        """
        if not EXCEL_AVAILABLE:
            raise ImportError(
                "openpyxl not installed. Run: pip install openpyxl"
            )

        self.config = config or ExcelConfig()
        self._workbook = Workbook()
        self._current_sheet = None
        self._sheet_row_counts = {}

        # Remove default sheet
        default_sheet = self._workbook.active
        if default_sheet:
            self._workbook.remove(default_sheet)

        self._setup_styles()
        app_logger.info("ExcelGenerator initialized")

    def _setup_styles(self) -> None:
        """Setup named styles."""
        # Header style
        header_style = NamedStyle(name="header_style")
        header_style.font = Font(
            name=self.config.font_family,
            size=self.config.font_size,
            bold=True,
            color=self.config.header_fg_color
        )
        header_style.fill = PatternFill(
            start_color=self.config.header_bg_color,
            end_color=self.config.header_bg_color,
            fill_type="solid"
        )
        header_style.alignment = Alignment(
            horizontal="right" if self.config.rtl else "left",
            vertical="center",
            wrap_text=True
        )
        header_style.border = Border(
            left=Side(style='thin', color=self.config.border_color),
            right=Side(style='thin', color=self.config.border_color),
            top=Side(style='thin', color=self.config.border_color),
            bottom=Side(style='thin', color=self.config.border_color)
        )

        # Data style
        data_style = NamedStyle(name="data_style")
        data_style.font = Font(
            name=self.config.font_family,
            size=self.config.font_size
        )
        data_style.alignment = Alignment(
            horizontal="right" if self.config.rtl else "left",
            vertical="center"
        )
        data_style.border = Border(
            left=Side(style='thin', color=self.config.border_color),
            right=Side(style='thin', color=self.config.border_color),
            top=Side(style='thin', color=self.config.border_color),
            bottom=Side(style='thin', color=self.config.border_color)
        )

        # Alt row style
        alt_style = NamedStyle(name="alt_row_style")
        alt_style.font = Font(
            name=self.config.font_family,
            size=self.config.font_size
        )
        alt_style.fill = PatternFill(
            start_color=self.config.alt_row_color,
            end_color=self.config.alt_row_color,
            fill_type="solid"
        )
        alt_style.alignment = Alignment(
            horizontal="right" if self.config.rtl else "left",
            vertical="center"
        )
        alt_style.border = Border(
            left=Side(style='thin', color=self.config.border_color),
            right=Side(style='thin', color=self.config.border_color),
            top=Side(style='thin', color=self.config.border_color),
            bottom=Side(style='thin', color=self.config.border_color)
        )

        # Number style
        number_style = NamedStyle(name="number_style")
        number_style.font = Font(
            name=self.config.font_family,
            size=self.config.font_size
        )
        number_style.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        number_style.number_format = '#,##0.00'

        # Currency style
        currency_style = NamedStyle(name="currency_style")
        currency_style.font = Font(
            name=self.config.font_family,
            size=self.config.font_size
        )
        currency_style.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        currency_style.number_format = '#,##0.00 "ر.س"'

        # Register styles
        try:
            self._workbook.add_named_style(header_style)
            self._workbook.add_named_style(data_style)
            self._workbook.add_named_style(alt_style)
            self._workbook.add_named_style(number_style)
            self._workbook.add_named_style(currency_style)
        except ValueError:
            # Styles already exist
            pass

    def add_sheet(
        self,
        name: str,
        data: List[Dict] = None,
        headers: List[str] = None,
        start_row: int = 1,
        start_col: int = 1
    ) -> 'ExcelGenerator':
        """
        Add a new sheet with data.

        Args:
            name: Sheet name
            data: Data rows as list of dicts
            headers: Column headers
            start_row: Starting row (1-indexed)
            start_col: Starting column (1-indexed)

        Returns:
            Self for chaining
        """
        # Create sheet
        sheet = self._workbook.create_sheet(title=name)
        self._current_sheet = sheet
        self._sheet_row_counts[name] = start_row

        # RTL direction
        if self.config.rtl:
            sheet.sheet_view.rightToLeft = True

        if not data:
            return self

        # Get headers from first row if not provided
        if headers is None:
            headers = list(data[0].keys())

        # Write headers
        for col_idx, header in enumerate(headers, start=start_col):
            cell = sheet.cell(row=start_row, column=col_idx, value=header)
            cell.style = "header_style"

        # Write data
        for row_idx, row_data in enumerate(data, start=start_row + 1):
            is_alt = (row_idx - start_row) % 2 == 0
            style_name = "alt_row_style" if is_alt else "data_style"

            for col_idx, header in enumerate(headers, start=start_col):
                value = row_data.get(header, "")
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)

                # Apply appropriate style based on value type
                if isinstance(value, (int, float)):
                    cell.style = "number_style"
                    if is_alt:
                        cell.fill = PatternFill(
                            start_color=self.config.alt_row_color,
                            end_color=self.config.alt_row_color,
                            fill_type="solid"
                        )
                else:
                    cell.style = style_name

        # Update row count
        self._sheet_row_counts[name] = start_row + len(data)

        # Auto-adjust column widths
        if self.config.auto_width:
            self._auto_adjust_columns(sheet, headers, start_col)

        # Freeze header row
        if self.config.freeze_header:
            sheet.freeze_panes = f'A{start_row + 1}'

        # Add auto-filter
        if self.config.auto_filter and data:
            last_col = get_column_letter(start_col + len(headers) - 1)
            sheet.auto_filter.ref = f"A{start_row}:{last_col}{start_row + len(data)}"

        return self

    def _auto_adjust_columns(
        self,
        sheet,
        headers: List[str],
        start_col: int
    ) -> None:
        """Auto-adjust column widths based on content."""
        for col_idx, header in enumerate(headers, start=start_col):
            column_letter = get_column_letter(col_idx)

            # Calculate max width
            max_width = len(str(header))

            for row in sheet.iter_rows(
                min_col=col_idx,
                max_col=col_idx,
                min_row=2,
                max_row=sheet.max_row
            ):
                for cell in row:
                    if cell.value:
                        max_width = max(max_width, len(str(cell.value)))

            # Set width with some padding
            sheet.column_dimensions[column_letter].width = min(max_width + 4, 50)

    def add_title(
        self,
        title: str,
        sheet_name: str = None,
        row: int = None
    ) -> 'ExcelGenerator':
        """
        Add title to sheet.

        Args:
            title: Title text
            sheet_name: Sheet name (uses current if not specified)
            row: Row number

        Returns:
            Self for chaining
        """
        sheet = self._get_sheet(sheet_name)
        if sheet is None:
            return self

        if row is None:
            row = 1

        cell = sheet.cell(row=row, column=1, value=title)
        cell.font = Font(
            name=self.config.font_family,
            size=16,
            bold=True,
            color=self.config.header_bg_color
        )
        cell.alignment = Alignment(
            horizontal="center" if not self.config.rtl else "right"
        )

        # Merge cells for title
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)

        return self

    def add_totals_row(
        self,
        columns: List[str],
        sheet_name: str = None,
        label: str = "الإجمالي"
    ) -> 'ExcelGenerator':
        """
        Add totals row at the bottom.

        Args:
            columns: Columns to sum
            sheet_name: Sheet name
            label: Label for totals row

        Returns:
            Self for chaining
        """
        sheet = self._get_sheet(sheet_name)
        if sheet is None:
            return self

        # Get current row count
        current_name = sheet_name or self._current_sheet.title
        last_row = self._sheet_row_counts.get(current_name, 1) + 1

        # Add label
        cell = sheet.cell(row=last_row, column=1, value=label)
        cell.font = Font(bold=True, name=self.config.font_family)

        # Find column indices and add SUM formulas
        header_row = sheet[1]
        for col_idx, cell in enumerate(header_row, start=1):
            if cell.value in columns:
                sum_cell = sheet.cell(row=last_row, column=col_idx)
                col_letter = get_column_letter(col_idx)
                sum_cell.value = f"=SUM({col_letter}2:{col_letter}{last_row - 1})"
                sum_cell.font = Font(bold=True, name=self.config.font_family)
                sum_cell.style = "currency_style"

        return self

    def add_chart(
        self,
        sheet_name: str = None,
        chart_type: str = "bar",
        title: str = "",
        data_range: str = None,
        categories_range: str = None,
        position: str = "E2",
        width: int = 15,
        height: int = 10
    ) -> 'ExcelGenerator':
        """
        Add chart to sheet.

        Args:
            sheet_name: Sheet name
            chart_type: Chart type (bar, pie, line, area, doughnut)
            title: Chart title
            data_range: Data range (e.g., "B2:B10")
            categories_range: Categories range (e.g., "A2:A10")
            position: Chart position (e.g., "E2")
            width: Chart width
            height: Chart height

        Returns:
            Self for chaining
        """
        sheet = self._get_sheet(sheet_name)
        if sheet is None:
            return self

        # Create chart based on type
        chart_classes = {
            "bar": BarChart,
            "pie": PieChart,
            "line": LineChart,
            "area": AreaChart,
            "doughnut": DoughnutChart
        }

        chart_class = chart_classes.get(chart_type.lower(), BarChart)
        chart = chart_class()

        chart.title = title
        chart.style = 10
        chart.width = width
        chart.height = height

        # Parse data range and add data
        if data_range:
            # Simple range parsing (e.g., "B2:B10")
            parts = data_range.replace(":", "").split()
            if len(parts) >= 2:
                data = Reference(sheet, range_string=data_range)
                chart.add_data(data, titles_from_data=True)

        if categories_range:
            cats = Reference(sheet, range_string=categories_range)
            chart.set_categories(cats)

        # Add data labels for pie charts
        if chart_type.lower() in ("pie", "doughnut"):
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showPercent = True
            chart.dataLabels.showVal = False

        sheet.add_chart(chart, position)
        return self

    def add_conditional_formatting(
        self,
        sheet_name: str = None,
        cell_range: str = "A1:Z100",
        rule_type: str = "color_scale",
        **kwargs
    ) -> 'ExcelGenerator':
        """
        Add conditional formatting.

        Args:
            sheet_name: Sheet name
            cell_range: Range to apply formatting
            rule_type: Type of formatting (color_scale, highlight, formula)
            **kwargs: Additional rule parameters

        Returns:
            Self for chaining
        """
        sheet = self._get_sheet(sheet_name)
        if sheet is None:
            return self

        if rule_type == "color_scale":
            rule = ColorScaleRule(
                start_type='min',
                start_color='F8696B',
                mid_type='percentile',
                mid_value=50,
                mid_color='FFEB84',
                end_type='max',
                end_color='63BE7B'
            )
            sheet.conditional_formatting.add(cell_range, rule)

        elif rule_type == "highlight":
            operator = kwargs.get('operator', 'greaterThan')
            value = kwargs.get('value', 0)
            fill = PatternFill(
                start_color=kwargs.get('color', 'FFFF00'),
                end_color=kwargs.get('color', 'FFFF00'),
                fill_type='solid'
            )
            rule = CellIsRule(
                operator=operator,
                formula=[str(value)],
                fill=fill
            )
            sheet.conditional_formatting.add(cell_range, rule)

        return self

    def add_data_validation(
        self,
        sheet_name: str = None,
        cell_range: str = "A1",
        validation_type: str = "list",
        values: List[str] = None,
        **kwargs
    ) -> 'ExcelGenerator':
        """
        Add data validation.

        Args:
            sheet_name: Sheet name
            cell_range: Range to apply validation
            validation_type: Type of validation (list, whole, decimal)
            values: List of allowed values
            **kwargs: Additional validation parameters

        Returns:
            Self for chaining
        """
        sheet = self._get_sheet(sheet_name)
        if sheet is None:
            return self

        if validation_type == "list" and values:
            dv = DataValidation(
                type="list",
                formula1=f'"{",".join(values)}"',
                showDropDown=False
            )
            dv.prompt = kwargs.get('prompt', 'اختر من القائمة')
            dv.promptTitle = kwargs.get('prompt_title', 'اختيار')
            sheet.add_data_validation(dv)
            dv.add(cell_range)

        return self

    def _get_sheet(self, sheet_name: str = None):
        """Get sheet by name or current sheet."""
        if sheet_name:
            return self._workbook[sheet_name] if sheet_name in self._workbook.sheetnames else None
        return self._current_sheet

    def protect_sheet(
        self,
        sheet_name: str = None,
        password: str = None
    ) -> 'ExcelGenerator':
        """
        Protect sheet with password.

        Args:
            sheet_name: Sheet name
            password: Protection password

        Returns:
            Self for chaining
        """
        sheet = self._get_sheet(sheet_name)
        if sheet is None:
            return self

        sheet.protection.sheet = True
        if password:
            sheet.protection.password = password

        return self

    def save(self, output_path: str) -> bool:
        """
        Save Excel file.

        Args:
            output_path: Output file path

        Returns:
            True if successful
        """
        try:
            # Ensure directory exists
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)

            # Set document properties
            self._workbook.properties.title = self.config.title
            self._workbook.properties.creator = self.config.author
            self._workbook.properties.company = self.config.company
            self._workbook.properties.subject = self.config.subject

            self._workbook.save(output_path)
            app_logger.info(f"Excel saved: {output_path}")
            return True

        except Exception as e:
            app_logger.error(f"Failed to save Excel: {e}", exc_info=True)
            return False

    def to_bytes(self) -> bytes:
        """
        Generate Excel as bytes.

        Returns:
            Excel content as bytes
        """
        try:
            buffer = io.BytesIO()
            self._workbook.save(buffer)
            return buffer.getvalue()

        except Exception as e:
            app_logger.error(f"Failed to generate Excel bytes: {e}")
            return b''

    def get_workbook(self) -> 'Workbook':
        """Get the underlying Workbook object."""
        return self._workbook


def create_excel_report(
    data: List[Dict],
    output_path: str,
    title: str = "",
    sheet_name: str = "البيانات",
    headers: List[str] = None,
    **config_kwargs
) -> bool:
    """
    Quick function to create Excel report.

    Args:
        data: Report data
        output_path: Output file path
        title: Report title
        sheet_name: Sheet name
        headers: Column headers
        **config_kwargs: Additional ExcelConfig options

    Returns:
        True if successful
    """
    try:
        config = ExcelConfig(title=title, **config_kwargs)
        excel = ExcelGenerator(config)

        # Add title if provided
        if title:
            excel.add_sheet(sheet_name, start_row=3)
            excel.add_title(title, row=1)
            # Re-add data starting from row 3
            sheet = excel._workbook[sheet_name]

            if headers is None and data:
                headers = list(data[0].keys())

            # Write headers at row 3
            for col_idx, header in enumerate(headers, start=1):
                cell = sheet.cell(row=3, column=col_idx, value=header)
                cell.style = "header_style"

            # Write data starting from row 4
            for row_idx, row_data in enumerate(data, start=4):
                for col_idx, header in enumerate(headers, start=1):
                    value = row_data.get(header, "")
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
        else:
            excel.add_sheet(sheet_name, data, headers=headers)

        return excel.save(output_path)

    except Exception as e:
        app_logger.error(f"Failed to create Excel report: {e}")
        return False
